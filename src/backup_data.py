#!/usr/bin/env python
# Copyright (c) 2023-2024 Berlin zaehlt Mobilitaet
# SPDX-License-Identifier: MIT

# @file    backup_data.py
# @author  Michael Behrisch
# @date    2023-01-03

import bisect
import csv
import datetime
import gzip
import json
import os
import sys
import zoneinfo

import openpyxl
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session

from common import ConnectionProvider, get_options
from datamodel import Base, TrafficCount, TrafficCountAdvanced, Segment, Camera


def open_session(options):
    engine = create_engine(options.database, echo=options.verbose > 1, future=True)
    Base.metadata.create_all(engine)
    return Session(engine)


def get_segments(session, options):
    segments = {}
    if options.json_file and os.path.exists(options.json_file):
        with open(options.json_file, encoding="utf8") as of:
            segment_data = json.load(of)
    else:
        for s in session.execute(select(Segment)):
            segments[s.Segment.id] = s.Segment
        return segments
    for segment in segment_data.get("features", []):
        sid = segment["properties"]["segment_id"]
        s = session.get(Segment, sid)
        if s is None:
            s = Segment(segment["properties"])
        else:
            s.update(segment["properties"])
        segments[sid] = s
    return segments


def get_cameras(session, conns, segments):
    cameras = conns.request("/v1/cameras", required="cameras")
    # cameras = json.load(open('cameras.json'))
    for camera in cameras.get("cameras", []):
        if camera["segment_id"] in segments:
            c = session.get(Camera, camera["instance_id"])
            if c is None:
                segments[camera["segment_id"]].add_camera(camera)


def update_db(segments, session, options, conns):
    print("Retrieving data for %s segments" % len(segments))
    newest_data = None
    for s in segments.values():
        session.add(s)
        active = [c.first_data_utc for c in s.cameras if c.first_data_utc is not None]
        if not active:
            print("No active camera for segment %s." % s.id, file=sys.stderr)
            continue
        first = s.last_backup_utc if s.last_backup_utc else min(active)
        last = s.last_data_utc
        if options.verbose and last is not None and first < last:
            print("Retrieving data for segment %s between %s and %s." % (s.id, first, last))
        while last is not None and first < last:
            if options.advanced:
                interval_end = first + datetime.timedelta(days=20)
                # TODO we can now specify the columns we need
                payload = '{"level": "segments", "format": "per-quarter", "id": "%s", "time_start": "%s", "time_end": "%s"}' % (s.id, first, interval_end)
                res = conns.request("/advanced/reports/traffic", "POST", payload, options.retry, "report")
                if res.get("status_code") == 403:
                    print(" Skipping %s." % s.id, file=sys.stderr)
                    break
            else:
                interval_end = first + datetime.timedelta(days=90)
                payload = '{"level": "segments", "format": "per-hour", "id": "%s", "time_start": "%s", "time_end": "%s"}' % (s.id, first, interval_end)
                res = conns.request("/v1/reports/traffic", "POST", payload, options.retry, "report")
            # if res.get("report", []):
            #     print(res)
            #     exit()
            for entry in res.get("report", []):
                if entry["uptime"] > 0:
                    tc = TrafficCountAdvanced(entry) if options.advanced else TrafficCount(entry)
                    idx = bisect.bisect(s.counts, tc.date_utc, key=lambda t: t.date_utc)
                    if not s.counts or s.counts[idx-1].date_utc != tc.date_utc:
                        s.counts.insert(idx, tc)
                    else:
                        s.counts[idx-1] = tc
            first = interval_end
        s.last_backup_utc = datetime.datetime.now(datetime.timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
        if last is not None and (newest_data is None or newest_data < last):
            newest_data = last
        session.commit()
    return newest_data


def round_separator(v, digits, sep):
    if sep:
        return str(round(v, digits)).replace(".", sep) if v is not None else ""
    return round(v, digits) if v is not None else None


def get_column_names():
    res = ["segment_id", "date_local", "uptime"]
    for mode in ("ped", "bike", "car", "heavy"):
        res += [mode + "_lft", mode + "_rgt", mode + "_total"]
    return res + ["v85"] + ["car_speed%s" % s for s in range(0, 80, 10)]


def get_column_values(tc, local_date, sep=None):
    result = [tc.segment_id, str(local_date)[:-9], round_separator(tc.uptime_rel, 6, sep)]
    for mode in ("pedestrian", "bike", "car", "heavy"):
        lft = getattr(tc, mode + "_lft")
        rgt = getattr(tc, mode + "_rgt")
        result += [round(lft) if lft is not None else None,
                   round(rgt) if rgt is not None else None,
                   round((lft or 0) + (rgt or 0)) if lft is not None or rgt is not None else None]
    result += [round_separator(tc.v85, 1, sep)]
    for v in tc.get_histogram():
        result.append(round_separator(v, 2, sep))
    return result


def write_xl(filename, segments, month=None):
    wb = openpyxl.Workbook()
    row = 1
    for s in segments:
        tzinfo=zoneinfo.ZoneInfo(s.timezone)
        for tc in s.counts:
            local_date = tc.date_utc.astimezone(tzinfo)
            if month is None or (local_date.year, local_date.month) == month:
                if row == 1:
                    for col, val in enumerate(get_column_names(), start=1):
                        wb.active.cell(row=row, column=col).value = val
                    row += 1
                for col, val in enumerate(get_column_values(tc, local_date), start=1):
                    wb.active.cell(row=row, column=col).value = val
                row += 1
    if row > 1:
        wb.save(filename)


def write_csv(filename, segments, month=None, delimiter=","):
    with gzip.open(filename, "wt") as csv_file:
        csv_out = csv.writer(csv_file, delimiter=delimiter)
        need_header = True
        for s in segments:
            tzinfo=zoneinfo.ZoneInfo(s.timezone)
            for tc in s.counts:
                local_date = tc.date_utc.astimezone(tzinfo)
                if month is None or (local_date.year, local_date.month) == month:
                    if need_header:
                        csv_out.writerow(get_column_names())
                        need_header = False
                    csv_out.writerow(get_column_values(tc, local_date))
    if need_header:  # no data
        os.remove(csv_file.name)


def add_month(offset, year, month):
    month += offset
    while month > 12:
        year += 1
        month -= 12
    while month < 1:
        year -= 1
        month += 12
    return year, month


def main(args=None):
    options = get_options(args)
    session = open_session(options)
    conns = ConnectionProvider(options.secrets["tokens"], options.url) if options.url else None
    excel = False
    segments = get_segments(session, options)
    if conns:
        get_cameras(session, conns, segments)
        session.commit()
        newest_data = update_db(segments, session, options, conns)
    else:
        newest_data = datetime.datetime.now(datetime.timezone.utc)
    if options.csv:
        if os.path.dirname(options.csv):
            os.makedirs(os.path.dirname(options.csv), exist_ok=True)
        curr_month = (newest_data.year, newest_data.month)
        month = (options.csv_start_year, 1) if options.csv_start_year else add_month(-1, *curr_month)
        while month <= curr_month:
            if excel:
                write_xl(options.csv + "_%s_%02i.xlsx" % month, segments.values(), month)
            else:
                write_csv(options.csv + "_%s_%02i.csv.gz" % month, segments.values(), month)
            month = add_month(1, *month)

    if options.csv_segments:
        if os.path.dirname(options.csv_segments):
            os.makedirs(os.path.dirname(options.csv_segments), exist_ok=True)
        for s in segments.values():
            if excel:
                write_xl(options.csv_segments + "_%s.csv.gz" % s.id, [s])
            else:
                write_csv(options.csv_segments + "_%s.csv.gz" % s.id, [s])

    if conns and options.verbose:
        conns.print_stats()


if __name__ == "__main__":
    main()
