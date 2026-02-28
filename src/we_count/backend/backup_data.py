#!/usr/bin/env python
# Copyright (c) 2023-2024 Berlin zaehlt Mobilitaet
# SPDX-License-Identifier: MIT

# @file    backup_data.py
# @author  Michael Behrisch
# @date    2023-01-03

import csv
import datetime
import gzip
import json
import os
import sys
import zoneinfo

import openpyxl
import pandas as pd

from common import ConnectionProvider, get_options, add_month, parse_utc


KEEP_COLUMNS = [
    "instance_id", "segment_id", "date", "uptime",
    "direction", "v85", "car_speed_hist_0to120plus",
    "heavy_lft", "heavy_rgt", "car_lft", "car_rgt",
    "bike_lft", "bike_rgt", "pedestrian_lft", "pedestrian_rgt",
    # advanced only:
    "mode_bus_lft", "mode_bus_rgt",
    "mode_lighttruck_lft", "mode_lighttruck_rgt",
    "mode_motorcycle_lft", "mode_motorcycle_rgt",
    "mode_stroller_lft", "mode_stroller_rgt",
    "mode_tractor_lft", "mode_tractor_rgt",
    "mode_trailer_lft", "mode_trailer_rgt",
    "mode_truck_lft", "mode_truck_rgt",
    "mode_night_lft", "mode_night_rgt",
    "speed_hist_car_lft", "speed_hist_car_rgt",
    "brightness", "sharpness",
]

def load_segments(json_file):
    segments = {}
    with open(json_file, encoding="utf8") as segment_file:
        for segment in json.load(segment_file).get("features", []):
            sid = segment["properties"]["segment_id"]
            segments[sid] = segment["properties"]
    return segments


def save_segments(segments, json_file):
    with open(json_file, encoding="utf8") as segment_file:
        content = json.load(segment_file)
    for segment in content.get("features", []):
        sid = segment["properties"]["segment_id"]
        if sid in segments:
            segment["properties"] = segments[sid]
    with open(json_file + ".new", "w", encoding="utf8") as segment_file:
        json.dump(content, segment_file, indent=2)
    os.rename(json_file + ".new", json_file)


def update_data(segments, df: pd.DataFrame, options, conns):
    print("Retrieving data for %s segments" % len(segments))
    newest_data = None
    backup_date = "last_advanced_backup" if options.advanced else "last_data_backup"
    for s in segments.values():
        active = [i["first_data_package"] for i in s["instance_ids"].values() if i["first_data_package"] is not None]
        if not active:
            print("No active camera for segment %s." % s["segment_id"], file=sys.stderr)
            continue
        first = parse_utc(s.get(backup_date, min(active)))
        last = parse_utc(s["last_data_package"])
        if options.verbose and last is not None and first < last:
            print("Retrieving data for segment %s between %s and %s." % (s["segment_id"], first, last))
        while last is not None and first < last:
            interval_end = first + datetime.timedelta(days=20)
            payload = {
                "level": "segments", "format": "per-hour", "id": s["segment_id"],
                "time_start": first.isoformat(), "time_end": interval_end.isoformat()}
            if options.advanced:
                payload.update(format="per-quarter", columns=",".join(KEEP_COLUMNS))
                res = conns.request("/advanced/reports/traffic", "POST", str(payload), options.retry, "report")
                if res.get("status_code") == 403:
                    print(" Skipping %s." % s["segment_id"], file=sys.stderr)
                    break
            else:
                res = conns.request("/v1/reports/traffic", "POST", str(payload), options.retry, "report")
            if options.dump:
                with open(options.dump, "a", encoding="utf8") as dump:
                    json.dump(res, dump, indent=2)
            report = res.get("report", [])
            if report:
                new_rows = pd.DataFrame(report)
                new_rows = new_rows[[c for c in KEEP_COLUMNS if c in new_rows.columns]]
                if df is None:
                    df = new_rows
                else:
                    df = pd.concat([df[~df["date"].isin(new_rows["date"]) | (df["segment_id"] != s["segment_id"])],
                                    new_rows], ignore_index=True)
            first = interval_end
        s[backup_date] = datetime.datetime.now(datetime.timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
        if last is not None and (newest_data is None or newest_data < last):
            newest_data = last
    return df, newest_data


def round_separator(v, digits, sep):
    if sep:
        return str(round(v, digits)).replace(".", sep) if v is not None else ""
    return round(v, digits) if v is not None else None


def get_column_names(advanced):
    res = ["segment_id", "date_local", "uptime"]
    for mode in TrafficCountAdvanced.modes() if advanced else TrafficCount.modes():
        if mode == "pedestrian":
            mode = "ped"
        res += [mode + "_lft", mode + "_rgt", mode + "_total"]
    return res + ["v85"] + ["car_speed%s" % s for s in range(0, 80, 10)]


def get_column_values(tc, local_date, advanced, sep=None):
    result = [tc.segment_id, str(local_date)[:-9], round_separator(tc.uptime_rel, 6, sep)]
    for mode in TrafficCountAdvanced.modes() if advanced else TrafficCount.modes():
        lft = getattr(tc, mode + "_lft")
        rgt = getattr(tc, mode + "_rgt")
        result += [round(lft) if lft is not None else None,
                   round(rgt) if rgt is not None else None,
                   round((lft or 0) + (rgt or 0)) if lft is not None or rgt is not None else None]
    result += [round_separator(tc.v85, 1, sep)]
    for v in tc.get_histogram():
        result.append(round_separator(v, 2, sep))
    return result


def _write_xl(filename, segments, df, advanced, month=None):
    wb = openpyxl.Workbook()
    row = 1
    for s in segments:
        tzinfo=zoneinfo.ZoneInfo(s["timezone"])
        for tc in s.counts:
            local_date = tc.date_utc.astimezone(tzinfo)
            if month is None or (local_date.year, local_date.month) == month:
                if row == 1:
                    for col, val in enumerate(get_column_names(advanced), start=1):
                        wb.active.cell(row=row, column=col).value = val
                    row += 1
                for col, val in enumerate(get_column_values(tc, local_date, advanced), start=1):
                    wb.active.cell(row=row, column=col).value = val
                row += 1
    if row > 1:
        wb.save(filename)


def _write_csv(filename, segments, df, advanced, month=None, delimiter=","):
    with gzip.open(filename, "wt") as csv_file:
        csv_out = csv.writer(csv_file, delimiter=delimiter)
        need_header = True
        for s in segments:
            tzinfo=zoneinfo.ZoneInfo(s["timezone"])
            for tc in s.counts:
                local_date = tc.date_utc.astimezone(tzinfo)
                if month is None or (local_date.year, local_date.month) == month:
                    if need_header:
                        csv_out.writerow(get_column_names(advanced))
                        need_header = False
                    csv_out.writerow(get_column_values(tc, local_date, advanced))
    if need_header:  # no data
        os.remove(csv_file.name)


def main(args=None):
    options = get_options(args)
    parquet_file = options.secrets["parquet_advanced"] if options.advanced else options.secrets["parquet"]
    df = pd.read_parquet(parquet_file) if os.path.exists(parquet_file) else None
    conns = ConnectionProvider(options.secrets["tokens"], options.url) if options.url else None
    excel = False
    segments = load_segments(options.json_file)
    if options.segments:
        filtered = [int(s.strip()) for s in options.segments.split(",")]
        filtered_segments = {k:v for k,v in segments.items() if k in filtered}
    else:
        filtered_segments = segments
    if conns:
        df, newest_data = update_data(filtered_segments, df, options, conns)
        if df is None:
            print("No data.", file=sys.stderr)
        else:
            df.to_parquet(parquet_file, index=False)
    else:
        newest_data = datetime.datetime.now(datetime.timezone.utc)
    save_segments(segments, options.json_file)
    if options.csv:
        if os.path.dirname(options.csv):
            os.makedirs(os.path.dirname(options.csv), exist_ok=True)
        curr_month = (newest_data.year, newest_data.month)
        month = (options.csv_start_year, 1) if options.csv_start_year else add_month(-1, *curr_month)
        while month <= curr_month:
            if excel:
                _write_xl(options.csv + "_%s_%02i.xlsx" % month, filtered_segments.values(), df, options.advanced, month)
            else:
                _write_csv(options.csv + "_%s_%02i.csv.gz" % month, filtered_segments.values(), df, options.advanced, month)
            month = add_month(1, *month)

    if options.csv_segments:
        if os.path.dirname(options.csv_segments):
            os.makedirs(os.path.dirname(options.csv_segments), exist_ok=True)
        for s in filtered_segments.values():
            if excel:
                _write_xl(options.csv_segments + "_%s.xlsx" % s.id, [s], options.advanced)
            else:
                _write_csv(options.csv_segments + "_%s.csv.gz" % s.id, [s], options.advanced)

    if conns and options.verbose:
        conns.print_stats()


if __name__ == "__main__":
    main()
